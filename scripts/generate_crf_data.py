from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "业务调研2（2026.）省儿童PICU调研-脓毒症病历.xlsx"
DATA_DIR = ROOT / "src" / "data"
RAW_DATA_DIR = DATA_DIR / "raw"
LEGACY_OUTPUT = ROOT / "src" / "data.js"


SOURCE_PATTERNS = {
    "EMR/住院病历": r"EMR|病历|病程|住院|出院|门诊|既往史|家族史|诊断|医嘱",
    "LIS/检验": r"LIS|lis|检验|血常规|生化|血气|凝血|培养|病原|尿常规",
    "护理/监护": r"护理|监护|生命体征|入院时间|心率|血压",
    "检查/PACS": r"检查信息|仪器检查|影像|心电图|超声|CT|MRI|X光|X线",
    "评分表": r"评分表|PSS|PIM3|昏迷评分|Glassgow|Glasgow",
    "随访": r"随访",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def slug(text: str, fallback: str) -> str:
    raw = re.sub(r"\s+", "_", text.strip())
    raw = re.sub(r"[^\w\u4e00-\u9fff]+", "_", raw).strip("_")
    return raw[:42] or fallback


def infer_source_systems(*parts: str) -> list[str]:
    text = " ".join(parts)
    systems = []
    for name, pattern in SOURCE_PATTERNS.items():
        if re.search(pattern, text, re.I):
            systems.append(name)
    return systems or ["来源待确认"]


def infer_control(field: str, options: list[str]) -> str:
    joined = " ".join([field, *options])
    if re.search(r"日期|时间", joined):
        return "date"
    if re.search(r"数值|mmHg|μ?mo?l|mmol|mmoL|×10|mg/L|s\b|=|计数|评分|心率|血压", joined):
        return "number"
    if len(options) <= 2 and any(option in {"无", "有", "是", "否"} or "○是" in option for option in options):
        return "boolean"
    if options:
        return "select"
    return "text"


def normalize_input_mode(modes: list[str], systems: list[str]) -> str:
    text = " ".join(modes)
    if "手动输入（无法提取）" in text:
        return "manual_unextractable"
    if "手动输入" in text or "需要手选" in text:
        return "manual"
    if "手动分类" in text:
        return "review"
    if "自动提取" in text or systems != ["来源待确认"]:
        return "auto"
    return "unknown"


def status_for(field: dict, field_index: int, case_index: int) -> str:
    mode = field["inputMode"]
    if mode in {"manual", "manual_unextractable"}:
        return "manual_required"
    if mode == "review":
        return "review_required"
    if mode == "unknown":
        return "source_unclear"
    cycle = (field_index + case_index) % 9
    if cycle in {0, 5}:
        return "review_required"
    if cycle == 7:
        return "missing"
    return "auto_filled"


def sample_value(field: dict, status: str, field_index: int) -> str:
    if status in {"missing", "source_unclear"}:
        return ""
    options = field["options"]
    if field["control"] == "number":
        return str([36.8, 7.2, 92, 118, 0.8, 12.4, 145, 23.1][field_index % 8])
    if field["control"] == "date":
        return ["2026-05-03", "2026-05-05", "2026-05-08"][field_index % 3]
    if options:
        return options[field_index % len(options)]
    return ["已记录", "未见异常", "待复核"][field_index % 3]


def build_template() -> dict:
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=False)
    sheet = workbook.active

    merged = {}
    for merged_range in sheet.merged_cells.ranges:
        value = clean(sheet.cell(merged_range.min_row, merged_range.min_col).value)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged[(row, col)] = value

    headers = [clean(sheet.cell(1, col).value) or f"col{col}" for col in range(1, sheet.max_column + 1)]
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        row = {}
        for col_idx, header in enumerate(headers, start=1):
            value = clean(sheet.cell(row_idx, col_idx).value) or merged.get((row_idx, col_idx), "")
            row[header] = value
        if row.get("模块") or row.get("字段") or row.get("选项"):
            rows.append(row)

    module_order = []
    module_rows = defaultdict(list)
    for row in rows:
        module = row["模块"].replace("\n", " ")
        if module not in module_order:
            module_order.append(module)
        module_rows[module].append(row)

    modules = []
    all_fields = []
    field_counter = 0
    for module_idx, module_name in enumerate(module_order, start=1):
        field_groups = defaultdict(list)
        for row in module_rows[module_name]:
            field_label = row["字段"] or row["选项"] or "未命名字段"
            field_groups[field_label].append(row)

        fields = []
        for field_label, grouped_rows in field_groups.items():
            field_counter += 1
            options = []
            sources = []
            roots = []
            notes = []
            input_modes = []
            annotations = []
            for row in grouped_rows:
                if row["选项"] and row["选项"] not in options:
                    options.append(row["选项"])
                if row["数据来源"] and row["数据来源"] not in sources:
                    sources.append(row["数据来源"])
                if row["数据根源"] and row["数据根源"] not in roots:
                    roots.append(row["数据根源"])
                if row["备注"] and row["备注"] not in notes:
                    notes.append(row["备注"])
                if row["是否需要输入"]:
                    input_modes.append(row["是否需要输入"])
                if row["是否标注"]:
                    annotations.append(row["是否标注"])

            systems = infer_source_systems(" ".join(sources), " ".join(roots), " ".join(notes), field_label)
            field = {
                "id": f"f{field_counter:03d}_{slug(module_name, 'module')}_{slug(field_label, 'field')}",
                "moduleId": f"m{module_idx:02d}_{slug(module_name, 'module')}",
                "module": module_name,
                "label": field_label,
                "options": options,
                "dataSource": "；".join(sources),
                "rootSource": "；".join(roots),
                "inputMode": normalize_input_mode(input_modes, systems),
                "rawInputMode": "；".join(dict.fromkeys(input_modes)),
                "annotationRequired": bool(annotations),
                "rawAnnotation": "；".join(dict.fromkeys(annotations)),
                "control": infer_control(field_label, options),
                "notes": "；".join(notes),
                "sourceSystems": systems,
            }
            fields.append(field)
            all_fields.append(field)

        module = {
            "id": f"m{module_idx:02d}_{slug(module_name, 'module')}",
            "name": module_name,
            "fieldCount": len(fields),
            "sourceSystems": sorted({system for field in fields for system in field["sourceSystems"]}),
            "fields": fields,
        }
        modules.append(module)

    return {
        "id": "picu-sepsis-v1",
        "name": "省儿童 PICU 脓毒症 CRF",
        "sourceFile": WORKBOOK.name,
        "moduleCount": len(modules),
        "fieldCount": len(all_fields),
        "sourceSystemCounts": Counter(system for field in all_fields for system in field["sourceSystems"]),
        "modules": modules,
        "fields": all_fields,
    }


def build_cases(fields: list[dict]) -> list[dict]:
    case_specs = [
        ("PICU-2026-0503", "7岁 男", "感染性休克 / 肺部感染", "李医生", "2026-05-03 09:22"),
        ("PICU-2026-0511", "2岁 女", "脓毒症 / 急性呼吸衰竭", "王医生", "2026-05-11 14:05"),
        ("PICU-2026-0520", "9月 男", "重症肺炎 / 免疫抑制待排", "陈医生", "2026-05-20 18:40"),
        ("PICU-2026-0528", "12岁 女", "腹腔感染 / 多器官功能障碍", "周医生", "2026-05-28 11:16"),
    ]
    cases = []
    for case_idx, (case_id, demographics, diagnosis, owner, updated_at) in enumerate(case_specs):
        values = {}
        status_counts = Counter()
        for field_idx, field in enumerate(fields):
            status = status_for(field, field_idx, case_idx)
            status_counts[status] += 1
            values[field["id"]] = {
                "value": sample_value(field, status, field_idx),
                "status": status,
                "confirmedBy": owner if status == "auto_filled" else "",
                "updatedAt": updated_at,
            }
        done = status_counts["auto_filled"] + status_counts["review_required"]
        completion = round(done / max(1, len(fields)) * 100)
        cases.append(
            {
                "id": case_id,
                "bed": f"PICU-{case_idx + 3:02d}",
                "demographics": demographics,
                "diagnosis": diagnosis,
                "owner": owner,
                "updatedAt": updated_at,
                "completion": completion,
                "statusCounts": dict(status_counts),
                "values": values,
            }
        )
    return cases


def build_evidence(fields: list[dict], cases: list[dict]) -> list[dict]:
    source_docs = {
        "EMR/住院病历": ("住院病历/诊断记录", "入院诊断提示感染相关诊断，出院诊断包含脓毒症及器官功能障碍描述。"),
        "LIS/检验": ("LIS 检验结果", "诊断后首个检验结果已回传，包含血常规、血气、生化、凝血及感染指标。"),
        "护理/监护": ("重症护理记录单", "入 PICU 后生命体征持续记录，系统保留血压、心率、呼吸、体温和尿量时间点。"),
        "检查/PACS": ("检查报告", "影像学及心电图检查报告可用于确认感染部位、心功能和呼吸相关信息。"),
        "评分表": ("PSS/PIM3 评分表", "评分表来源于病历表单，部分项目可从结构化记录回填，部分需要医生确认。"),
        "随访": ("出院后随访", "放弃治疗后结局需要随访补充，院内系统仅能提供自动出院或转院线索。"),
        "来源待确认": ("来源待确认", "原始调研表未明确稳定来源，一期标记为待治理字段。"),
    }
    evidence = []
    for case in cases:
        for system, (title, snippet) in source_docs.items():
            related = [field["id"] for field in fields if system in field["sourceSystems"]][:12]
            if not related:
                continue
            evidence.append(
                {
                    "id": f"{case['id']}-{slug(system, 'source')}",
                    "caseId": case["id"],
                    "system": system,
                    "title": title,
                    "time": case["updatedAt"],
                    "snippet": snippet,
                    "relatedFields": related,
                }
            )
    return evidence


def fields_by_system(fields: list[dict], system: str, limit: int = 8) -> list[str]:
    return [field["id"] for field in fields if system in field["sourceSystems"]][:limit]


def build_raw_tables(fields: list[dict], cases: list[dict]) -> list[dict]:
    lis_fields = fields_by_system(fields, "LIS/检验", 10)
    emr_fields = fields_by_system(fields, "EMR/住院病历", 10)
    nursing_fields = fields_by_system(fields, "护理/监护", 8)
    exam_fields = fields_by_system(fields, "检查/PACS", 8)
    score_fields = fields_by_system(fields, "评分表", 8)
    followup_fields = fields_by_system(fields, "随访", 6)

    raw = []
    for idx, case in enumerate(cases):
        raw.append(
            {
                "caseId": case["id"],
                "tables": [
                    {
                        "id": "patient_profile",
                        "name": "患者基本信息表",
                        "system": "EMR/住院病历",
                        "columns": ["字段", "值", "来源"],
                        "rows": [
                            {"字段": "病例号", "值": case["id"], "来源": "住院首页"},
                            {"字段": "床位", "值": case["bed"], "来源": "PICU床位表"},
                            {"字段": "年龄性别", "值": case["demographics"], "来源": "患者主索引"},
                            {"字段": "主要诊断", "值": case["diagnosis"], "来源": "入院/出院诊断"},
                        ],
                        "linkedFields": emr_fields[:4],
                    },
                    {
                        "id": "diagnosis_orders",
                        "name": "诊断与医嘱表",
                        "system": "EMR/住院病历",
                        "columns": ["时间", "类型", "名称", "状态"],
                        "rows": [
                            {"时间": case["updatedAt"], "类型": "诊断", "名称": case["diagnosis"], "状态": "已归档"},
                            {"时间": case["updatedAt"], "类型": "医嘱", "名称": "抗感染治疗", "状态": "执行中"},
                            {"时间": case["updatedAt"], "类型": "医嘱", "名称": "PICU重症监护", "状态": "执行中"},
                        ],
                        "linkedFields": emr_fields,
                    },
                    {
                        "id": "lis_results",
                        "name": "LIS检验结果表",
                        "system": "LIS/检验",
                        "columns": ["采样时间", "项目", "结果", "单位", "异常"],
                        "rows": [
                            {"采样时间": case["updatedAt"], "项目": "白细胞计数(WBC)", "结果": str(13.2 + idx), "单位": "×10^9/L", "异常": "高"},
                            {"采样时间": case["updatedAt"], "项目": "CRP", "结果": str(88 + idx * 7), "单位": "mg/L", "异常": "高"},
                            {"采样时间": case["updatedAt"], "项目": "PCT", "结果": str(round(8.4 + idx * 0.8, 1)), "单位": "ng/mL", "异常": "高"},
                            {"采样时间": case["updatedAt"], "项目": "乳酸", "结果": str(round(3.1 + idx * 0.3, 1)), "单位": "mmol/L", "异常": "高"},
                        ],
                        "linkedFields": lis_fields,
                    },
                    {
                        "id": "nursing_vitals",
                        "name": "护理生命体征表",
                        "system": "护理/监护",
                        "columns": ["记录时间", "体温", "心率", "呼吸", "血压", "SpO2"],
                        "rows": [
                            {"记录时间": case["updatedAt"], "体温": "38.7", "心率": str(142 + idx * 4), "呼吸": str(34 + idx), "血压": "82/46", "SpO2": "92%"},
                            {"记录时间": case["updatedAt"], "体温": "37.9", "心率": str(130 + idx * 3), "呼吸": str(30 + idx), "血压": "91/52", "SpO2": "95%"},
                        ],
                        "linkedFields": nursing_fields,
                    },
                    {
                        "id": "exam_reports",
                        "name": "检查报告表",
                        "system": "检查/PACS",
                        "columns": ["检查时间", "检查类型", "结论"],
                        "rows": [
                            {"检查时间": case["updatedAt"], "检查类型": "胸部影像", "结论": "双肺感染性改变，建议结合临床及实验室指标。"},
                            {"检查时间": case["updatedAt"], "检查类型": "心电图", "结论": "窦性心动过速，PR/QRS间期待结构化回填。"},
                        ],
                        "linkedFields": exam_fields,
                    },
                    {
                        "id": "score_forms",
                        "name": "PSS/PIM3评分表",
                        "system": "评分表",
                        "columns": ["评分时间", "评分项", "记录值", "备注"],
                        "rows": [
                            {"评分时间": case["updatedAt"], "评分项": "PIM3-收缩压", "记录值": "82 mmHg", "备注": "入ICU 1小时内最低值"},
                            {"评分时间": case["updatedAt"], "评分项": "PSS-循环", "记录值": "异常", "备注": "需医生确认"},
                            {"评分时间": case["updatedAt"], "评分项": "PSS-呼吸", "记录值": "异常", "备注": "结合血气和机械通气记录"},
                        ],
                        "linkedFields": score_fields,
                    },
                    {
                        "id": "followup",
                        "name": "随访记录表",
                        "system": "随访",
                        "columns": ["随访日期", "随访方式", "结局", "备注"],
                        "rows": [
                            {"随访日期": "2026-06-01", "随访方式": "电话", "结局": "待确认", "备注": "放弃治疗后结局需补充"},
                        ],
                        "linkedFields": followup_fields,
                    },
                ],
            }
        )
    return raw


RAW_TABLE_EXPORTS = {
    "patient_profile": ("patient-profile", "patientProfile"),
    "diagnosis_orders": ("diagnosis-orders", "diagnosisOrders"),
    "lis_results": ("lis-results", "lisResults"),
    "nursing_vitals": ("nursing-vitals", "nursingVitals"),
    "exam_reports": ("exam-reports", "examReports"),
    "score_forms": ("score-forms", "scoreForms"),
    "followup": ("followup", "followup"),
}


def ts_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def write_ts(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body, encoding="utf-8")


def write_data_files(template: dict, cases: list[dict], evidence: list[dict], raw_tables: list[dict]) -> None:
    template = {**template, "sourceSystemCounts": dict(template["sourceSystemCounts"])}

    write_ts(
        DATA_DIR / "crf-template.ts",
        'import type { CrfTemplate } from "@/types";\n\n'
        f"export const crfTemplate = {ts_json(template)} satisfies CrfTemplate;\n",
    )
    write_ts(
        DATA_DIR / "case-records.ts",
        'import type { CaseRecord } from "@/types";\n\n'
        f"export const caseRecords = {ts_json(cases)} satisfies CaseRecord[];\n",
    )
    write_ts(
        DATA_DIR / "source-evidence.ts",
        'import type { SourceEvidence } from "@/types";\n\n'
        f"export const sourceEvidence = {ts_json(evidence)} satisfies SourceEvidence[];\n",
    )
    write_ts(
        DATA_DIR / "status-labels.ts",
        'import type { InputModeLabels, StatusLabels } from "@/types";\n\n'
        'export const statusLabels = {\n'
        '  "auto_filled": "已自动带入",\n'
        '  "manual_required": "需手填",\n'
        '  "review_required": "需确认",\n'
        '  "missing": "缺失",\n'
        '  "source_unclear": "来源不明确"\n'
        "} satisfies StatusLabels;\n\n"
        "export const inputModeLabels = {\n"
        '  "auto": "自动提取",\n'
        '  "manual": "手动输入",\n'
        '  "manual_unextractable": "手动输入（无法提取）",\n'
        '  "review": "自动提取后人工分类",\n'
        '  "unknown": "待确认"\n'
        "} satisfies InputModeLabels;\n",
    )

    flattened: dict[str, list[dict]] = {table_id: [] for table_id in RAW_TABLE_EXPORTS}
    for raw_set in raw_tables:
        for table in raw_set["tables"]:
            flattened[table["id"]].append({"caseId": raw_set["caseId"], **table})

    raw_imports = []
    raw_exports = []
    for table_id, (filename, export_name) in RAW_TABLE_EXPORTS.items():
        write_ts(
            RAW_DATA_DIR / f"{filename}.ts",
            'import type { RawTable } from "@/types";\n\n'
            f"export const {export_name} = {ts_json(flattened[table_id])} satisfies RawTable[];\n",
        )
        raw_imports.append(f'import {{ {export_name} }} from "./raw/{filename}";')
        raw_exports.append(export_name)

    write_ts(
        DATA_DIR / "index.ts",
        'export { crfTemplate } from "./crf-template";\n'
        'export { caseRecords } from "./case-records";\n'
        'export { sourceEvidence } from "./source-evidence";\n'
        'export { inputModeLabels, statusLabels } from "./status-labels";\n'
        + "\n".join(raw_imports)
        + "\n\n"
        + f"export const rawTables = [{', '.join(raw_exports)}].flat();\n",
    )

    if LEGACY_OUTPUT.exists():
        LEGACY_OUTPUT.unlink()


def main() -> None:
    template = build_template()
    cases = build_cases(template["fields"])
    evidence = build_evidence(template["fields"], cases)
    raw_tables = build_raw_tables(template["fields"], cases)
    write_data_files(template, cases, evidence, raw_tables)
    print(f"Generated {DATA_DIR.relative_to(ROOT)}")
    print(f"Modules: {template['moduleCount']}; fields: {template['fieldCount']}; evidence: {len(evidence)}")


if __name__ == "__main__":
    main()
